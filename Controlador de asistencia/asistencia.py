import cv2
import face_recognition as fr
import os
import numpy
from datetime import datetime
import openpyxl

# crear base de datos
ruta = 'Empleados'
mis_imagenes = []
nombres_empleados = []
lista_empleados = os.listdir(ruta)

for nombre in lista_empleados:
    imagen_actual = cv2.imread(f'{ruta}/{nombre}')
    mis_imagenes.append(imagen_actual)
    nombres_empleados.append(os.path.splitext(nombre)[0])

print(nombres_empleados)

# codificar imagenes
def codificar(imagenes):
    # crear una lista nueva
    lista_codificada = []

    # pasar todas las imagenes a rgb
    for imagen in imagenes:
        imagen = cv2.cvtColor(imagen, cv2.COLOR_BGR2RGB)

        # codificar
        codificado = fr.face_encodings(imagen)[0]

        # agregar a la lista
        lista_codificada.append(codificado)

    # devolver lista codificada
    return lista_codificada

# registro de ingresos
def registrar_ingresos(persona):
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    hora_actual = datetime.now().strftime("%H:%M:%S")
    registro_existente = False

    wb = openpyxl.load_workbook("asistencia.xlsx")
    sheet = wb.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == persona and row[2] == "hora de entrada":
            sheet.cell(row=sheet.max_row, column=3).value = f"{fecha_actual} {hora_actual}"
            registro_existente = True
            break

    if not registro_existente:
        sheet.append([persona, f"{fecha_actual} {hora_actual}", "hora de entrada"])

    wb.save("asistencia.xlsx")

    if registro_existente:
        print(f"Hora de salida registrada para {persona}.")
    else:
        print(f"Hora de entrada registrada para {persona}.")

lista_empleados_codificada = codificar(mis_imagenes)

# tomar imagen de la camara
captura = cv2.VideoCapture(0, cv2.CAP_DSHOW)

# leer la imagen de la camara
exito, imagen = captura.read()

if not exito:
    print("No ha sido posible realizar la caprtura")
else:
    # reconocer cara en captura
    cara_captura = fr.face_locations(imagen)

    # codificar la cara capturada
    cara_captura_codificada = fr.face_encodings(imagen, cara_captura)

    # buscar coincidencias
    for caracodif, caraubic in zip(cara_captura_codificada, cara_captura):
        coincidencias = fr.compare_faces(lista_empleados_codificada, caracodif)
        distancias = fr.face_distance(lista_empleados_codificada, caracodif)

        print(distancias)

        indice_coincidencia = numpy.argmin(distancias)

        #mostrar coincidencias si las hay
        if distancias[indice_coincidencia] > 0.6:
            print("No coincide con ningún empleado")
        else:
            
            # buscar el nombre del empleado encontrado
            nombre = nombres_empleados[indice_coincidencia]

            # mostrar rectangulo
            y1, x2, y2, x1 = caraubic
            cv2.rectangle(imagen, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.rectangle(imagen, (x1, y2 -35), (x2, y2), (0, 255, 0), cv2.FILLED)
            cv2.putText(imagen, nombre, (x1 + 6, y2 -6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

            registrar_ingresos(nombre)

            # mostrar la imagen obtenida
            cv2.imshow("Imagen Web", imagen)

            # mantener abierta la ventana
            cv2.waitKey(0)